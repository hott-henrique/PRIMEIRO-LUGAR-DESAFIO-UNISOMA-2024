import collections
import datetime
import json
import logging

import numpy as np

import openpyxl
import openpyxl.styles
import openpyxl.chart
import openpyxl.chart.label
import openpyxl.chart.legend
import openpyxl.chart.text

class NumpyEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, np.ndarray):
            return "NUMPY ARRAY - TOO BIG TO PRINT - EDIT THE CODE TO SEE."
            return str(o.tolist())
        return super().default(o)

def clear_past_output(file_path: str):
    workbook = openpyxl.load_workbook(file_path)

    if "Inconsistência" in workbook.sheetnames:
        sheet = workbook["Inconsistência"]
        sheet.delete_rows(idx=2, amount=sheet.max_row - 1)

    if "Agendamento" in workbook.sheetnames:
        sheet = workbook["Agendamento"]
        sheet.delete_rows(idx=2, amount=sheet.max_row - 1)

    if "Análise" in workbook.sheetnames:
        sheet = workbook["Análise"]
        sheet.delete_rows(idx=2, amount=sheet.max_row - 1)

    if "Solução" in workbook.sheetnames:
        sheet = workbook["Solução"]
        sheet.delete_rows(idx=2, amount=sheet.max_row - 1)

    workbook.save(file_path)

def save_errors(file_path: str, errors: list, logger: logging.Logger = logging.getLogger()):
    logger.debug(f"errors: {errors}")

    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook['Inconsistência']

    max_row = sheet.max_row

    for row in range(max_row, 1, -1):
        sheet.delete_rows(row)

    for error in errors:
        sheet.append(( error["table"], error["type"], error["message"], datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") ))

    workbook.save(file_path)

def save_output(file_path: str, response: dict, solution: np.ndarray, logger: logging.Logger = logging.getLogger()):
    solve_columns = list()

    days_map = { i: day for i, day in enumerate([ "seg", "ter", "qua", "qui", "sex", "sab" ]) }
    hour_map = { i: hour for i, hour in enumerate([ "hr_" + str(x + 8) for x in range(13) ]) }

    logger.debug(json.dumps(response, indent=2, ensure_ascii=False, skipkeys=True, cls=NumpyEncoder))

    for row in solution:
        new_row = list()

        # ===================================================
        new_row.append(response['patient_names'][row[1]])
        new_row.append(response['doctor_names'][row[0]])
        # ===================================================
        new_row.append(days_map[row[2]])
        new_row.append(hour_map[row[3]])
        new_row.append(response["local_names"][row[4]])
        new_row.append(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        solve_columns.append(new_row)

    logger.debug(solve_columns)

    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook['Solução']

    max_row = sheet.max_row

    for row in range(max_row, 1, -1):
        sheet.delete_rows(row)

    for solve in solve_columns:
        sheet.append(solve)

    map_local = response["local_names"]
    map_professionals = response["doctor_names"]
    map_patient = response['patient_names']
    map_day = [ "seg", "ter", "qua", "qui", "sex", "sab" ]

    create_schedulling(workbook, solution, map_local, map_patient, map_professionals)

    if "Análise" in workbook.sheetnames:
        workbook.remove(workbook["Análise"])

    kpi_sheet = workbook.create_sheet(title="Análise")

    professionals_per_location = [ (str(map_local[i]), int(v)) for i, v in enumerate(response["local_m_l"].sum(axis=0)) ]

    appointments_per_location = [
        (str(map_local[row[0]]), int(row[1]))
        for row in np.column_stack(np.unique(solution[:, 4], return_counts=True))
    ]

    appointments_per_day = [
        (str(map_day[row[0]]), int(row[1]))
        for row in np.column_stack(np.unique(solution[:, 2], return_counts=True))
    ]

    appointments_per_professional = [
        (str(map_professionals[row[0]]), int(row[1]))
        for row in np.column_stack(np.unique(solution[:, 0], return_counts=True))
    ]

    base_row = 2

    plot(kpi_sheet, "F5",
         professionals_per_location, base_row=base_row,
         title="Profissionais por Localidade",
         x_axis_title="Local",
         y_axis_title="# Profissionais")

    base_row = base_row + len(professionals_per_location) + 1

    plot(kpi_sheet, "F20",
         appointments_per_location, base_row=base_row,
         title="Agendamentos por Localidade",
         x_axis_title="Local",
         y_axis_title="# Agendamentos")

    base_row = base_row + len(appointments_per_location) + 1

    plot(kpi_sheet, "O5",
         appointments_per_day,
         base_row=base_row,
         title="Agendamentos por Dia da Semana",
         x_axis_title="Dia da Semana",
         y_axis_title="# Agendamentos")

    base_row = base_row + len(appointments_per_day) + 1

    plot(kpi_sheet, "O20",
         appointments_per_professional,
         base_row=base_row,
         title="Agendamentos por Profissional",
         x_axis_title="Profisionnal",
         y_axis_title="# Agendamentos")

    workbook.save(file_path)

def create_schedulling(
        workbook: openpyxl.Workbook,
        solution: np.ndarray,
        map_local: dict[int, str],
        map_patient: dict[int, str],
        map_professionals: dict[int, str]
    ):
    if "Agendamento" in workbook.sheetnames:
        workbook.remove(workbook["Agendamento"])

    sheet = workbook.create_sheet(title="Agendamento")

    thin = openpyxl.styles.Side(border_style="thin", color="000000")

    per_doctor_schedulle = collections.defaultdict(list)

    for row in solution:
        per_doctor_schedulle[int(row[0])].append(row)

    base_row, base_column = 2, 2

    for k, apointemnts in per_doctor_schedulle.items():
        # INFO: Title cells.
        sheet.merge_cells(None, start_row=base_row, end_row=base_row, start_column=base_column, end_column=base_column + 6)

        sheet.cell(row=base_row, column=base_column).alignment = openpyxl.styles.Alignment(horizontal='center')

        # INFO: Time of the day.
        for i in range(13):
            sheet.cell(row=base_row + 1 + (i + 1), column=base_column).value = 8 + i

        for i, day in enumerate([ "seg", "ter", "qua", "qui", "sex", "sab" ]):
            sheet.cell(row=base_row + 1, column=base_column + i + 1).value = day

        local = 0
        for m, p, d, h, l in apointemnts:
            patient_name = map_patient[p]
            patient_text = f"{patient_name}*" if l == 0 else patient_name

            local = l if l != 0 else local

            sheet.cell(row=base_row + 1 + (h + 1), column=base_column + (d + 1)).value = patient_text

        for r in range(base_row, base_row + 15):
            for c in range(base_column, base_column + 7):
                sheet.cell(r, c).border = openpyxl.styles.Border(top=thin, bottom=thin, left=thin, right=thin)

        sheet.cell(row=base_row, column=base_column).value = f"{map_professionals[k]} - {map_local[local]}"

        base_row += 13 + 3

def plot(sheet, cell: str, data: list[tuple[str, int]], base_row: int, title: str, x_axis_title: str, y_axis_title: str):
    sheet.append([ "", "" ])

    for row in data:
        sheet.append(row)

    data_reference = openpyxl.chart.Reference(sheet,
                                              min_col=1,
                                              max_col=2,
                                              min_row=base_row,
                                              max_row=base_row + len(data))

    chart = openpyxl.chart.BarChart()

    chart.add_data(data_reference, from_rows=True, titles_from_data=True)

    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title

    sheet.add_chart(chart, cell)
